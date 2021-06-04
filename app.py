from flask import Flask,render_template,request,url_for,redirect,flash,send_file
import xlrd
from openpyxl import load_workbook

app=Flask(__name__)


@app.route("/")
@app.route("/home")
def index():
    return render_template("index.html")


#sem marks
@app.route("/sem",methods=["POST"])
def sem():
    if request.method == "POST":
        loc=request.form.get("loc")
        mail=request.form.get("email")
        name=request.form.get("name")
        path=loc
        wb=xlrd.open_workbook(path)
        ws=wb.sheet_by_index(0)
        row=(ws.nrows)
        col=(ws.ncols)
        data=[[ws.cell_value(r,c)for c in range(col)]for r in range(row)]
 
        roll= request.form.get("id")
        roll=roll.upper()
        subcode=[]
        sub=[]
        grade=[]
        credit=[]
        count=0
        flag=0
        for i in range(len(data)):
    
            if roll==(data[i][0]):

                subcode.append(data[i][1])
                sub.append(data[i][2])
                grade.append(data[i][3])
                credit.append(data[i][4])
                count+=1
                if count==9:
                    break
                if count!=0:
                    flag=1
        
        mylist=zip(subcode,sub,grade,credit)
        #data insert
        fname="report.xlsx"
        wb=load_workbook(fname)

        ws=wb.worksheets[1]
        
        data =[name,mail,roll]
        ws.append(data)
        wb.save(fname)
        return render_template("result.html" ,mylist=mylist ,roll=roll,flag=flag)

    return render_template("index.html")



@app.route("/mid",methods=["POST"])
def mid():
    if request.method == "POST":
        loc=request.form.get("loc")
        mail=request.form.get("email")
        name=request.form.get("name")
        path=loc
        wb=xlrd.open_workbook(path)
        ws=wb.sheet_by_index(0)
        row=(ws.nrows)
        col=(ws.ncols)
        data=[[ws.cell_value(r,c)for c in range(col)]for r in range(row)]
 
        roll= request.form.get("id")
        roll=roll.upper()
        subcode=[]
        sub=[]
        grade=[]
        credit=[]
        count=0
        flag=0
        for i in range(len(data)):
    
            if roll==(data[i][0]):

                subcode.append(data[i][1])
                sub.append(data[i][2])
                grade.append(data[i][3])
                credit.append(data[i][4])
                count+=1
                if count==9:
                    break
                if count!=0:
                    flag=1
        
        mylist=zip(subcode,sub,grade,credit)
         #data insert
        fname="report.xlsx"
        wb=load_workbook(fname)

        ws=wb.worksheets[2]
        
        data =[name,mail,roll]
        ws.append(data)
        wb.save(fname)
        return render_template("result.html" ,mylist=mylist ,roll=roll,flag=flag)

    return render_template("index.html")


@app.route("/report",methods=["POST"])
def report():
    if request.method == "POST":
        name=request.form.get("name")
        email=request.form.get("email")
        sub=request.form.get("sub")
        mess=request.form.get("mess")
        fname="report.xlsx"
        wb=load_workbook(fname)

        ws=wb.worksheets[0]
        
        data =[name,email,sub,mess]
        ws.append(data)
        wb.save(fname) 
        
        return render_template("index.html")
    return render_template("index.html")

@app.route('/download')
def export_db():

    return send_file('report.xlsx')



if __name__ == '__main__':
    app.run(debug=True)